"""
relatorios.py
Geração de relatórios em PDF e Excel.
"""

import os
from datetime import datetime
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


RENDIMENTO_ESPERADO = 14500
FONTE = "Arial"


def _get_pasta_relatorios() -> str:
    import sys
    if getattr(sys, 'frozen', False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    pasta = os.path.join(base, "db", "relatorios_exportados")
    os.makedirs(pasta, exist_ok=True)
    return pasta


def _fmt_numero(valor) -> str:
    if isinstance(valor, (int, float)):
        return f"{valor:,.0f}".replace(",", ".")
    return str(valor)


def _fmt_moeda(valor) -> str:
    inteiro = int(valor)
    centavos = round((valor - inteiro) * 100)
    milhares = f"{inteiro:,}".replace(",", ".")
    return f"R$ {milhares},{centavos:02d}"


def _fmt_data(data_str) -> str:
    if not data_str:
        return ""
    if "-" in str(data_str):
        p = data_str.split("-")
        if len(p) == 3:
            return f"{p[2]}/{p[1]}/{p[0]}"
    return data_str


# ==================== PDF ====================

class PDF(FPDF):
    def __init__(self):
        super().__init__(orientation='L', format='A4')
        self.set_auto_page_break(auto=True, margin=15)

    def header(self):
        self.set_font(FONTE, 'B', 16)
        self.set_text_color(68, 114, 196)
        self.cell(0, 10, 'Relatorio de Rendimento de Toners', 0, 1, 'C')
        self.set_font(FONTE, '', 10)
        self.set_text_color(0, 0, 0)
        self.cell(0, 6, f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}', 0, 1, 'C')
        self.ln(3)

    def footer(self):
        self.set_y(-15)
        self.set_font(FONTE, 'I', 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f'Pagina {self.page_no()}', 0, 0, 'C')

    def titulo_maquina(self, nome, periodo):
        self.set_font(FONTE, 'B', 13)
        self.set_text_color(68, 114, 196)
        self.cell(0, 10, f'Maquina: {nome}', 0, 1, 'L')
        self.set_font(FONTE, '', 10)
        self.set_text_color(0, 0, 0)
        self.cell(0, 6, f'Periodo: {periodo}', 0, 1, 'L')
        self.ln(4)

    # Colunas: Cor(30) DataInst(30) DataRet(30) ContIni(35) ContFim(35) Impressoes(35) Custo(35) R$/pag(37) = 267mm
    COLS = [('Cor', 30), ('Data Inst.', 30), ('Data Ret.', 30),
            ('Cont. Inicial', 35), ('Cont. Final', 35),
            ('Impressoes', 35), ('Custo (R$)', 35), ('R$/pagina', 37)]

    def cabecalho_tabela(self):
        self.set_font(FONTE, 'B', 9)
        self.set_fill_color(68, 114, 196)
        self.set_text_color(255, 255, 255)
        for titulo, largura in self.COLS:
            self.cell(largura, 10, titulo, 1, 0, 'C', 1)
        self.ln()

    def linha_tabela(self, dados):
        self.set_font(FONTE, '', 8)
        em_uso = dados.get('em_uso', False)
        if em_uso:
            self.set_fill_color(200, 240, 210)
        elif dados.get('total_impressoes', 0) < RENDIMENTO_ESPERADO and not em_uso:
            self.set_fill_color(255, 200, 200)
        else:
            self.set_fill_color(255, 255, 255)
        self.set_text_color(0, 0, 0)

        valores = [
            dados['cor'],
            _fmt_data(dados['data_instalacao']),
            "EM USO" if em_uso else _fmt_data(dados.get('data_retirada', '')),
            _fmt_numero(dados['contador_inicial']),
            "-" if em_uso else _fmt_numero(dados.get('contador_final', 0)),
            "-" if em_uso else _fmt_numero(dados.get('total_impressoes', 0)),
            _fmt_moeda(dados['custo']),
            "-" if em_uso else f"{dados.get('custo_pagina', 0):.4f}".replace('.', ',')
        ]
        for i, ((_, larg), val) in enumerate(zip(self.COLS, valores)):
            self.cell(larg, 8, str(val), 1, 0, 'C' if i < 3 else 'R', 1)
        self.ln()

    def linha_total(self, total_paginas, total_custo, custo_medio, cont_ini=None, cont_fim=None):
        self.set_font(FONTE, 'B', 9)
        self.set_fill_color(217, 225, 242)
        self.set_text_color(0, 0, 0)
        self.cell(90, 10, 'TOTAIS', 1, 0, 'C', 1)
        self.cell(35, 10, _fmt_numero(cont_ini) if cont_ini else '-', 1, 0, 'R', 1)
        self.cell(35, 10, _fmt_numero(cont_fim) if cont_fim else '-', 1, 0, 'R', 1)
        self.cell(35, 10, _fmt_numero(total_paginas), 1, 0, 'R', 1)
        self.cell(35, 10, _fmt_moeda(total_custo), 1, 0, 'R', 1)
        self.cell(37, 10, f"{custo_medio:.4f}".replace('.', ','), 1, 1, 'R', 1)

    def estatisticas(self, rendimentos, toners_ativos):
        total = len(rendimentos) + len(toners_ativos)
        if total == 0:
            return
        self.ln(8)
        self.set_font(FONTE, 'B', 11)
        self.cell(0, 8, 'Analise de Rendimento', 0, 1, 'L')
        self.set_font(FONTE, '', 10)
        self.cell(60, 6, 'Total de toners:', 0, 0)
        self.cell(60, 6, f'{total} ({len(toners_ativos)} em uso)', 0, 1)

        toners_baixo = sum(1 for r in rendimentos if r.total_impressoes < RENDIMENTO_ESPERADO)
        if toners_baixo > 0:
            self.set_text_color(200, 0, 0)
            self.cell(60, 6, 'Toners abaixo da meta:', 0, 0)
            self.cell(60, 6, f'{toners_baixo}', 0, 1)
            self.set_text_color(0, 0, 0)

        if rendimentos:
            media = sum(r.total_impressoes for r in rendimentos) / len(rendimentos)
            self.cell(60, 6, 'Media de impressoes:', 0, 0)
            self.cell(60, 6, f'{_fmt_numero(media)} pags/toner', 0, 1)


def gerar_relatorio_pdf(rendimentos, maquina_nome, periodo,
                        caminho_destino=None, toners_ativos=None):
    if toners_ativos is None:
        toners_ativos = []

    pasta = _get_pasta_relatorios()
    pdf = PDF()
    pdf.add_page()
    pdf.titulo_maquina(maquina_nome, periodo)

    if rendimentos or toners_ativos:
        pdf.cabecalho_tabela()

        total_paginas = 0
        total_custo = 0
        cont_iniciais = []
        cont_finais = []

        for t in toners_ativos:
            pdf.linha_tabela({
                'cor': t.cor, 'data_instalacao': t.data_instalacao,
                'data_retirada': None, 'contador_inicial': t.contador_inicial,
                'contador_final': None, 'total_impressoes': 0,
                'custo': t.custo, 'custo_pagina': 0, 'em_uso': True
            })
            total_custo += t.custo
            cont_iniciais.append(t.contador_inicial)

        for r in rendimentos:
            pdf.linha_tabela({
                'cor': r.cor, 'data_instalacao': r.data_instalacao,
                'data_retirada': r.data_retirada, 'contador_inicial': r.contador_inicial,
                'contador_final': r.contador_final, 'total_impressoes': r.total_impressoes,
                'custo': r.custo, 'custo_pagina': r.custo_pagina, 'em_uso': False
            })
            total_paginas += r.total_impressoes
            total_custo += r.custo
            cont_iniciais.append(r.contador_inicial)
            if r.contador_final:
                cont_finais.append(r.contador_final)

        custo_medio = total_custo / total_paginas if total_paginas > 0 else 0
        pdf.ln(4)
        pdf.linha_total(total_paginas, total_custo, custo_medio,
                        min(cont_iniciais) if cont_iniciais else None,
                        max(cont_finais) if cont_finais else None)
        pdf.estatisticas(rendimentos, toners_ativos)
    else:
        pdf.set_font(FONTE, '', 12)
        pdf.cell(0, 10, 'Nenhum toner registrado no periodo selecionado.', 0, 1, 'C')

    if not caminho_destino:
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        caminho_destino = os.path.join(pasta, f"toners_{maquina_nome.replace(' ','_')}_{ts}.pdf")

    os.makedirs(os.path.dirname(caminho_destino) if os.path.dirname(caminho_destino) else '.', exist_ok=True)
    pdf.output(caminho_destino)
    return caminho_destino


# ==================== EXCEL ====================

def gerar_relatorio_excel(dados_maquinas, periodo, caminho_destino=None):
    pasta = _get_pasta_relatorios()
    wb = Workbook()
    wb.remove(wb.active)

    st_cab  = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    fi_cab  = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    st_tot  = Font(name='Calibri', size=10, bold=True)
    fi_tot  = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    st_alr  = Font(name='Calibri', size=10, bold=True, color='CC0000')
    fi_alr  = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
    fi_ati  = PatternFill(start_color='C8F0D2', end_color='C8F0D2', fill_type='solid')
    st_ati  = Font(name='Calibri', size=10, bold=True, color='1A6B3A')
    aln_ctr = Alignment(horizontal='center', vertical='center')

    for dado in dados_maquinas:
        nome_aba = dado['nome'][:31]
        ws = wb.create_sheet(title=nome_aba)

        ws.merge_cells('A1:H1')
        ws['A1'].value = f"Relatorio de Rendimento de Toners - {dado['nome']}"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True)
        ws['A1'].alignment = aln_ctr
        ws.row_dimensions[1].height = 25

        ws.merge_cells('A2:H2')
        ws['A2'].value = f"Periodo: {periodo or 'Todos'}"
        ws['A2'].font = Font(italic=True)
        ws['A2'].alignment = aln_ctr

        ws.merge_cells('A3:H3')
        ws['A3'].value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A3'].font = Font(italic=True)
        ws['A3'].alignment = aln_ctr

        cabecalhos = ['Cor', 'Data Instalacao', 'Data Retirada',
                      'Contador Inicial', 'Contador Final',
                      'Total Impressoes', 'Custo (R$)', 'R$/pagina']
        for col, cab in enumerate(cabecalhos, 1):
            c = ws.cell(row=5, column=col, value=cab)
            c.font = st_cab; c.fill = fi_cab; c.alignment = aln_ctr

        rendimentos = dado.get('rendimentos', [])
        ativos = dado.get('ativos', [])

        if rendimentos or ativos:
            linha = 6
            total_paginas = 0
            total_custo = 0
            cont_iniciais = []
            cont_finais = []

            for t in ativos:
                def _set(col, val, fmt=None, font=None, fill=None):
                    c = ws.cell(row=linha, column=col, value=val)
                    if fmt: c.number_format = fmt
                    if font: c.font = font
                    if fill: c.fill = fill
                    return c
                ws.cell(row=linha, column=1, value=t.cor).font = st_ati
                ws.cell(row=linha, column=1).fill = fi_ati
                ws.cell(row=linha, column=2, value=_fmt_data(t.data_instalacao)).fill = fi_ati
                c3 = ws.cell(row=linha, column=3, value="EM USO"); c3.font = st_ati; c3.fill = fi_ati
                c4 = ws.cell(row=linha, column=4, value=t.contador_inicial); c4.number_format='#,##0'; c4.fill=fi_ati
                ws.cell(row=linha, column=5, value="-").fill = fi_ati
                ws.cell(row=linha, column=6, value="-").fill = fi_ati
                c7 = ws.cell(row=linha, column=7, value=round(t.custo, 2)); c7.number_format='"R$" #,##0.00'; c7.fill=fi_ati
                ws.cell(row=linha, column=8, value="-").fill = fi_ati
                total_custo += t.custo
                cont_iniciais.append(t.contador_inicial)
                linha += 1

            for r in rendimentos:
                ws.cell(row=linha, column=1, value=r.cor)
                ws.cell(row=linha, column=2, value=_fmt_data(r.data_instalacao))
                ws.cell(row=linha, column=3, value=_fmt_data(r.data_retirada))
                c4 = ws.cell(row=linha, column=4, value=r.contador_inicial); c4.number_format='#,##0'
                c5 = ws.cell(row=linha, column=5, value=r.contador_final); c5.number_format='#,##0'
                c6 = ws.cell(row=linha, column=6, value=r.total_impressoes); c6.number_format='#,##0'
                if r.total_impressoes < RENDIMENTO_ESPERADO: c6.font=st_alr; c6.fill=fi_alr
                c7 = ws.cell(row=linha, column=7, value=round(r.custo, 2)); c7.number_format='"R$" #,##0.00'
                c8 = ws.cell(row=linha, column=8, value=round(r.custo_pagina, 4)); c8.number_format='#,##0.0000'
                total_paginas += r.total_impressoes
                total_custo += r.custo
                cont_iniciais.append(r.contador_inicial)
                if r.contador_final: cont_finais.append(r.contador_final)
                linha += 1

            lt = linha + 1
            ws.merge_cells(f'A{lt}:C{lt}')
            c = ws.cell(row=lt, column=1, value='TOTAIS'); c.font=st_tot; c.fill=fi_tot

            ci = ws.cell(row=lt, column=4, value=min(cont_iniciais) if cont_iniciais else None)
            ci.font=st_tot; ci.fill=fi_tot; ci.number_format='#,##0'

            cf = ws.cell(row=lt, column=5, value=max(cont_finais) if cont_finais else None)
            cf.font=st_tot; cf.fill=fi_tot; cf.number_format='#,##0'

            cp = ws.cell(row=lt, column=6, value=total_paginas)
            cp.font=st_tot; cp.fill=fi_tot; cp.number_format='#,##0'

            cc = ws.cell(row=lt, column=7, value=round(total_custo, 2))
            cc.font=st_tot; cc.fill=fi_tot; cc.number_format='"R$" #,##0.00'

            custo_medio = total_custo / total_paginas if total_paginas > 0 else 0
            cm = ws.cell(row=lt, column=8, value=round(custo_medio, 4))
            cm.font=st_tot; cm.fill=fi_tot; cm.number_format='#,##0.0000'

            # Resumo
            ls = lt + 3
            ws.merge_cells(f'A{ls}:C{ls}')
            ws.cell(row=ls, column=1, value='RESUMO').font = Font(bold=True, size=11)
            total_todos = len(rendimentos) + len(ativos)
            ws.cell(row=ls+1, column=1, value='Total de toners:')
            ws.cell(row=ls+1, column=2, value=f'{total_todos} toner(s) ({len(ativos)} em uso)')
            if rendimentos:
                media = sum(r.total_impressoes for r in rendimentos) / len(rendimentos)
                ws.cell(row=ls+2, column=1, value='Media de impressoes:')
                ws.cell(row=ls+2, column=2, value=f"{media:,.0f} paginas/toner".replace(",", "."))
            baixos = sum(1 for r in rendimentos if r.total_impressoes < RENDIMENTO_ESPERADO)
            if baixos > 0:
                ws.cell(row=ls+3, column=1, value='Toners abaixo da meta:')
                c = ws.cell(row=ls+3, column=2, value=f'{baixos} toner(s)')
                c.font = Font(color='CC0000')
        else:
            ws.merge_cells('A6:H6')
            ws['A6'].value = 'Nenhum toner registrado no periodo selecionado.'
            ws['A6'].font = Font(italic=True)
            ws['A6'].alignment = aln_ctr

        larguras = [12, 16, 16, 16, 16, 18, 18, 14]
        for col, larg in enumerate(larguras, 1):
            ws.column_dimensions[get_column_letter(col)].width = larg

    if not dados_maquinas:
        ws = wb.create_sheet(title='Sem dados')
        ws['A1'] = 'Nenhum dado encontrado.'

    if not caminho_destino:
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        if len(dados_maquinas) == 1:
            caminho_destino = os.path.join(pasta, f"toners_{dados_maquinas[0]['nome'].replace(' ','_')}_{ts}.xlsx")
        else:
            caminho_destino = os.path.join(pasta, f"relatorio_toners_{ts}.xlsx")

    os.makedirs(os.path.dirname(caminho_destino) if os.path.dirname(caminho_destino) else '.', exist_ok=True)
    wb.save(caminho_destino)
    return caminho_destino
